"""
Excel Servisi — Excel dosyaları oluşturma, okuma, düzenleme.
openpyxl kullanır.
"""

import json
import logging
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)


def excel_oku(yol: str, sayfa: str = None) -> str:
    """Excel dosyasını oku ve tablo formatında döndür."""
    try:
        import openpyxl
    except ImportError:
        return "❌ openpyxl kurulu değil. pip install openpyxl"

    try:
        path = Path(yol).expanduser().resolve()
        if not path.exists():
            return f"❌ Dosya bulunamadı: {path}"

        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

        if sayfa:
            if sayfa not in wb.sheetnames:
                return f"❌ '{sayfa}' sayfası bulunamadı. Mevcut: {', '.join(wb.sheetnames)}"
            ws = wb[sayfa]
        else:
            ws = wb.active

        cikti = f"📊 {path.name}"
        if len(wb.sheetnames) > 1:
            cikti += f" (Sayfalar: {', '.join(wb.sheetnames)})"
        cikti += f"\n📄 Aktif: {ws.title}\n\n"

        satirlar = []
        for row in ws.iter_rows(values_only=True):
            satirlar.append([str(cell) if cell is not None else "" for cell in row])

        if not satirlar:
            wb.close()
            return cikti + "Dosya boş."

        # Sütun genişliklerini hesapla
        sutun_genislikleri = []
        for col_idx in range(len(satirlar[0])):
            max_len = 0
            for satir in satirlar[:50]:  # İlk 50 satırdan hesapla
                if col_idx < len(satir):
                    max_len = max(max_len, len(satir[col_idx]))
            sutun_genislikleri.append(min(max_len, 25))

        # Tablo formatı
        max_satir = min(len(satirlar), 100)
        for i, satir in enumerate(satirlar[:max_satir]):
            hücreler = []
            for j, s in enumerate(satir):
                gen = sutun_genislikleri[j] if j < len(sutun_genislikleri) else 15
                hücreler.append(s[:gen].ljust(gen))
            cikti += " | ".join(hücreler) + "\n"
            if i == 0:
                cikti += "-+-".join("-" * g for g in sutun_genislikleri) + "\n"

        if len(satirlar) > 100:
            cikti += f"\n... ve {len(satirlar) - 100} satır daha"

        cikti += f"\n\nToplam: {len(satirlar)} satır, {len(satirlar[0]) if satirlar else 0} sütun"

        wb.close()
        return cikti

    except Exception as e:
        logger.error(f"Excel okuma hatası: {e}")
        return f"❌ Excel okuma hatası: {str(e)}"


def excel_olustur(yol: str, basliklar: list, veriler: list = None,
                  sayfa_adi: str = "Sayfa1") -> str:
    """Yeni Excel dosyası oluştur. Stillenmiş başlıklar + veri."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        return "❌ openpyxl kurulu değil. pip install openpyxl"

    try:
        path = Path(yol).expanduser().resolve()
        # .xlsx uzantısı yoksa ekle
        if path.suffix.lower() not in (".xlsx", ".xls"):
            path = path.with_suffix(".xlsx")
        path.parent.mkdir(parents=True, exist_ok=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sayfa_adi

        # Stil tanımları
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Başlıklar
        for col, baslik in enumerate(basliklar, 1):
            cell = ws.cell(row=1, column=col, value=baslik)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Veriler
        if veriler:
            for row_idx, satir in enumerate(veriler, 2):
                for col_idx, deger in enumerate(satir, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=deger)
                    cell.border = thin_border
                    # Sayısal değerleri dönüştürmeyi dene
                    if isinstance(deger, str):
                        try:
                            cell.value = float(deger)
                        except (ValueError, TypeError):
                            pass

        # Sütun genişliklerini otomatik ayarla
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                val_len = len(str(cell.value or ""))
                max_len = max(max_len, val_len)
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

        # Filtreleme ekle
        ws.auto_filter.ref = ws.dimensions

        wb.save(str(path))
        wb.close()

        satir_sayisi = len(veriler) if veriler else 0
        return (
            f"✅ Excel oluşturuldu: {path}\n"
            f"📊 {satir_sayisi} satır, {len(basliklar)} sütun"
        )

    except Exception as e:
        logger.error(f"Excel oluşturma hatası: {e}")
        return f"❌ Excel oluşturma hatası: {str(e)}"


def excel_duzenle(yol: str, islemler: list) -> str:
    """
    Excel dosyasını düzenle.
    islemler: [
        {"hucre": "A1", "deger": "yeni değer"},
        {"satir_ekle": ["val1", "val2", ...]},
        {"satir_sil": 5}  # 5. satırı sil
    ]
    """
    try:
        import openpyxl
    except ImportError:
        return "❌ openpyxl kurulu değil. pip install openpyxl"

    try:
        path = Path(yol).expanduser().resolve()
        if not path.exists():
            return f"❌ Dosya bulunamadı: {path}"

        wb = openpyxl.load_workbook(str(path))
        ws = wb.active

        degisiklik_sayisi = 0
        detaylar = []

        for islem in islemler:
            if "hucre" in islem and "deger" in islem:
                eski = ws[islem["hucre"]].value
                ws[islem["hucre"]] = islem["deger"]
                detaylar.append(f"  📝 {islem['hucre']}: {eski} → {islem['deger']}")
                degisiklik_sayisi += 1

            elif "satir_ekle" in islem:
                ws.append(islem["satir_ekle"])
                detaylar.append(f"  ➕ Satır eklendi: {islem['satir_ekle'][:3]}...")
                degisiklik_sayisi += 1

            elif "satir_sil" in islem:
                satir_no = islem["satir_sil"]
                ws.delete_rows(satir_no)
                detaylar.append(f"  🗑️ {satir_no}. satır silindi")
                degisiklik_sayisi += 1

        wb.save(str(path))
        wb.close()

        cikti = f"✅ Excel güncellendi: {path}\n📝 {degisiklik_sayisi} değişiklik:\n"
        cikti += "\n".join(detaylar[:10])
        if len(detaylar) > 10:
            cikti += f"\n  ... ve {len(detaylar) - 10} değişiklik daha"

        return cikti

    except Exception as e:
        logger.error(f"Excel düzenleme hatası: {e}")
        return f"❌ Excel düzenleme hatası: {str(e)}"
